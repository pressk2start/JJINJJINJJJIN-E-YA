# -*- coding: utf-8 -*-
"""
multichain_balance_checker.py  v5.0  — API 키 없이 동작
대화형: 체인 선택 → 주소 입력 → native + token 잔고 조회
소수점 끝자리까지 정확히 보존
지원 (18종):
  BTC          mempool.space (무료, 키 불필요)
  LTC          litecoinspace.org / Trezor Blockbook 폴백 (무료)
  BCH          Zelcore Blockbook (무료)
  RVN          Ravencoin Blockbook (무료)
  ETH          공개 RPC + Blockscout v2 토큰 (무료)
  BNB          공개 RPC + 주요토큰 batch RPC (무료)
  POL          공개 RPC + Blockscout v2 토큰 (무료)
  KAIA         공개 RPC + Kaiascan (키 있으면 토큰 조회)
  SGB          공개 RPC (무료)
  XRP          공식 Ripple RPC (무료)
  XLM          Horizon API (무료)
  SOL          공개 RPC (무료)
  TRX          TronGrid (무료, 키 있으면 rate limit ↑)
  ONT          공식 노드 다중 폴백 (무료)
  ARDR         Jelurida 공개 노드 (무료)
  IGNIS        Ardor child chain 2 (무료)
  EOS          Hyperion (무료)
  PCI          scan.payprotocol.io (무료, 비공식)
선택 환경변수 (없어도 동작, 있으면 rate limit 완화):
  TRONGRID_API_KEY, KAIASCAN_API_KEY, SOL_RPC_URL, XRP_RPC_URL, ARDOR_NODE_URL
"""
from __future__ import annotations
import os, sys, json, time, requests
from decimal import Decimal, InvalidOperation
from typing import Any, List, Optional
TIMEOUT = 20
S = requests.Session()
S.headers.update({"User-Agent": "multichain-balance-checker/5.0"})
# ═══════════════════════════════════════════════
#  공통 유틸
# ═══════════════════════════════════════════════
def http_get(url, headers=None, params=None) -> Any:
    r = S.get(url, headers=headers or {}, params=params or {}, timeout=TIMEOUT)
    r.raise_for_status()
    ct = r.headers.get("content-type", "")
    return r.json() if ("application/json" in ct or r.text[:1] in ("{", "[")) else r.text
def http_post(url, payload, headers=None) -> Any:
    r = S.post(url, json=payload, headers=headers or {}, timeout=TIMEOUT)
    r.raise_for_status()
    ct = r.headers.get("content-type", "")
    return r.json() if ("application/json" in ct or r.text[:1] in ("{", "[")) else r.text
def hex2int(h: str) -> str:
    return str(int(h, 16))
def fmt(raw: str | int, dec: int) -> str:
    """base-unit 정수 → 정확한 소수 문자열"""
    s = str(raw).strip()
    neg = s.startswith("-")
    if neg: s = s[1:]
    if not s.isdigit():
        raise ValueError(f"정수가 아님: {raw}")
    if dec == 0:
        out = s
    else:
        s = s.zfill(dec + 1)
        out = f"{s[:-dec]}.{s[-dec:]}"
    return f"-{out}" if neg else out
def noexp(v: str) -> str:
    t = str(v).strip()
    if not t: return "0"
    if "e" not in t.lower(): return t
    try: return format(Decimal(t), "f")
    except InvalidOperation: return t
def R(chain, addr, ticker, amount, raw=None, dec=None, ct=None, typ="token", **kw):
    d = dict(chain=chain, address=addr, ticker=ticker, amount=amount,
             raw_amount=raw, decimals=dec, contract=ct, type=typ)
    d.update(kw)
    return d
# ═══════════════════════════════════════════════
#  BTC  — mempool.space (키 불필요)
# ═══════════════════════════════════════════════
def get_btc(addr):
    d = http_get(f"https://mempool.space/api/address/{addr}")
    cs = d["chain_stats"]
    ms = d["mempool_stats"]
    funded = cs["funded_txo_sum"] + ms["funded_txo_sum"]
    spent  = cs["spent_txo_sum"]  + ms["spent_txo_sum"]
    raw = str(funded - spent)
    return [R("BTC", addr, "BTC", fmt(raw, 8), raw, 8, None, "native")]
# ═══════════════════════════════════════════════
#  LTC  — 다중 폴백 (키 불필요)
# ═══════════════════════════════════════════════
def get_ltc(addr):
    # litecoinspace.org (mempool 방식)
    try:
        d = http_get(f"https://litecoinspace.org/api/address/{addr}")
        cs = d["chain_stats"]
        ms = d["mempool_stats"]
        funded = cs["funded_txo_sum"] + ms["funded_txo_sum"]
        spent  = cs["spent_txo_sum"]  + ms["spent_txo_sum"]
        raw = str(funded - spent)
        return [R("LTC", addr, "LTC", fmt(raw, 8), raw, 8, None, "native")]
    except Exception:
        pass
    # Trezor Blockbook 폴백
    try:
        d = http_get(f"https://ltc1.trezor.io/api/v2/address/{addr}?details=basic")
        raw = str(d.get("balance", "0"))
        return [R("LTC", addr, "LTC", fmt(raw, 8), raw, 8, None, "native")]
    except Exception:
        pass
    # BlockCypher 최종 폴백
    d = http_get(f"https://api.blockcypher.com/v1/ltc/main/addrs/{addr}/balance")
    raw = str(d.get("balance", "0"))
    return [R("LTC", addr, "LTC", fmt(raw, 8), raw, 8, None, "native")]
# ═══════════════════════════════════════════════
#  BCH  — Zelcore Blockbook (키 불필요)
# ═══════════════════════════════════════════════
def get_bch(addr):
    d = http_get(f"https://blockbook.bch.zelcore.io/api/v2/address/{addr}?details=basic")
    raw = str(d.get("balance", "0"))
    return [R("BCH", addr, "BCH", fmt(raw, 8), raw, 8, None, "native")]
# ═══════════════════════════════════════════════
#  RVN  — Ravencoin Blockbook (키 불필요)
# ═══════════════════════════════════════════════
def get_rvn(addr):
    try:
        d = http_get(f"https://blockbook.ravencoin.org/api/v2/address/{addr}?details=basic")
        raw = str(d.get("balance", "0"))
        return [R("RVN", addr, "RVN", fmt(raw, 8), raw, 8, None, "native")]
    except Exception:
        bal = str(http_get(f"https://chainz.cryptoid.info/rvn/api.dws?q=getbalance&a={addr}")).strip()
        return [R("RVN", addr, "RVN", noexp(bal), None, 8, None, "native")]
# ═══════════════════════════════════════════════
#  EVM 공통: 공개 RPC (ETH / BNB / POL)
#  native = eth_getBalance
#  token  = Blockscout v2 (ETH, POL) / batch RPC (BNB)
# ═══════════════════════════════════════════════
EVM_CFG = {
    "ETH": {
        "rpc": "https://ethereum-rpc.publicnode.com",
        "blockscout": "https://eth.blockscout.com",
        "dec": 18,
    },
    "BNB": {
        "rpc": "https://bsc-dataseed.bnbchain.org",
        "blockscout": None,  # BSC Blockscout 미지원
        "dec": 18,
    },
    "POL": {
        "rpc": "https://polygon-bor-rpc.publicnode.com",
        "blockscout": "https://polygon.blockscout.com",
        "dec": 18,
    },
}
# BSC 주요 BEP20 토큰 (batch RPC용)
BSC_KNOWN_TOKENS = {
    "0x55d398326f99059fF775485246999027B3197955": ("USDT", 18),
    "0x8AC76a51cc950d9822D68b83fE1Ad97B32Cd580d": ("USDC", 18),
    "0xe9e7CEA3DedcA5984780Bafc599bD69ADd087D56": ("BUSD", 18),
    "0xbb4CdB9CBd36B01bD1cBaEBF2De08d9173bc095c": ("WBNB", 18),
    "0x2170Ed0880ac9A755fd29B2688956BD959F933F8": ("ETH", 18),
    "0x1D2F0da169ceB9fC7B3144628dB156f3F6c60dBE": ("XRP", 18),
    "0x7130d2A12B9BCbFAe4f2634d864A1Ee1Ce3Ead9c": ("BTCB", 18),
    "0x0E09FaBB73Bd3Ade0a17ECC321fD13a19e81cE82": ("CAKE", 18),
    "0x3EE2200Efb3400fAbB9AacF31297cBdD1d435D47": ("ADA", 18),
    "0xbA2aE424d960c26247Dd6c32edC70B295c744C43": ("DOGE", 8),
    "0x1CE0c2827e2eF14D5C4f29a091d735A204794041": ("AVAX", 18),
    "0xF8A0BF9cF54Bb92F17374d9e9A321E6a111a51bD": ("LINK", 18),
    "0x7083609fCE4d1d8Dc0C979AAb8c869Ea2C873402": ("DOT", 18),
    "0x4338665CBB7B2485A8855A139b75D5e34AB0DB94": ("LTC", 18),
    "0x8fF795a6F4D97E7887C79beA79aba5cc76444aDf": ("BCH", 18),
    "0x1Fa4a73a3F0133f0025378af00236f3aBDEE5D63": ("NEAR", 18),
    "0xCE7de646e7208a4Ef112cb6ed5038FA6cC6b12e3": ("TRX", 6),
    "0xCC42724C6683B7E57334c4E856f4c9965ED682bD": ("MATIC", 18),
    "0x570A5D26f7765Ecb712C0924E4De545B89fD43dF": ("SOL", 18),
    "0x76A797A59Ba2C17726896976B7B3747BfD1d220f": ("TON", 9),
}
def _evm_rpc(rpc_url, method, params):
    return http_post(rpc_url, {
        "jsonrpc": "2.0", "id": 1, "method": method, "params": params
    })
def _evm_batch_rpc(rpc_url, calls):
    """배치 RPC 요청"""
    r = S.post(rpc_url, json=calls, timeout=TIMEOUT)
    r.raise_for_status()
    return r.json()
def _blockscout_tokens(blockscout_url, addr):
    """Blockscout v2 API로 토큰 잔고 조회"""
    url = f"{blockscout_url}/api/v2/addresses/{addr}/token-balances"
    data = http_get(url)
    if not isinstance(data, list):
        return []
    results = []
    for item in data:
        token = item.get("token", {})
        value = str(item.get("value", "0"))
        if value == "0" or not value:
            continue
        sym = token.get("symbol") or token.get("name") or "UNKNOWN"
        dec = int(token.get("decimals") or 0)
        ct = token.get("address_hash") or token.get("address") or ""
        ttype = token.get("type", "ERC-20")
        if ttype not in ("ERC-20",):
            continue
        results.append((sym, value, dec, ct))
    return results
def _bsc_batch_token_check(rpc_url, addr):
    """BSC: 주요 토큰을 batch RPC로 조회"""
    addr_padded = "0" * 24 + addr[2:].lower()
    calls = []
    token_list = list(BSC_KNOWN_TOKENS.items())
    for i, (ca, (sym, dec)) in enumerate(token_list):
        calls.append({
            "jsonrpc": "2.0", "id": i + 1,
            "method": "eth_call",
            "params": [{"to": ca, "data": "0x70a08231" + addr_padded}, "latest"]
        })
    resp = _evm_batch_rpc(rpc_url, calls)
    results = []
    resp_map = {r["id"]: r for r in resp}
    for i, (ca, (sym, dec)) in enumerate(token_list):
        r = resp_map.get(i + 1, {})
        hex_val = r.get("result", "0x0")
        if hex_val and hex_val != "0x" and hex_val != "0x0":
            raw = str(int(hex_val, 16))
            if raw != "0":
                results.append((sym, raw, dec, ca))
    return results
def get_evm_balances(chain, addr):
    cfg = EVM_CFG[chain]
    # native
    resp = _evm_rpc(cfg["rpc"], "eth_getBalance", [addr, "latest"])
    raw = hex2int(resp["result"])
    results = [R(chain, addr, chain, fmt(raw, cfg["dec"]), raw, cfg["dec"], None, "native")]
    # tokens
    token_results = []
    try:
        if cfg.get("blockscout"):
            token_results = _blockscout_tokens(cfg["blockscout"], addr)
        elif chain == "BNB":
            token_results = _bsc_batch_token_check(cfg["rpc"], addr)
    except Exception:
        # BSC fallback: batch RPC
        if chain == "BNB":
            try:
                token_results = _bsc_batch_token_check(cfg["rpc"], addr)
            except Exception:
                pass
    for sym, raw_val, dec, ct in token_results:
        results.append(R(chain, addr, sym, fmt(raw_val, dec), raw_val, dec, ct))
    return results
# ═══════════════════════════════════════════════
#  KAIA  — 공개 RPC + Kaiascan (키 필수)
# ═══════════════════════════════════════════════
KAIA_RPC = "https://public-en.node.kaia.io"
# KAIA 주요 토큰 (batch RPC 폴백용)
KAIA_KNOWN_TOKENS = {
    "0xceE8FAF64bB97a73bb51E115Aa89C17FfA8dD167": ("oUSDT", 6),
    "0x754288077D0fF82AF7a5317C7CB8c444D421d103": ("oUSDC", 6),
    "0x5C74070FDeA071359b86082bd9f9b3dEaafbe32b": ("oKLAY", 18),
    "0x34d21b1e550D73cee41151c77F3c73359527a396": ("oETH", 18),
    "0x16D0e1fBD024c600Ca7BF8120Bd834E240F1801b": ("oWBTC", 8),
    "0x5096dB80B21Ef45230C9E423C373f1FC9C0198dd": ("WEMIX", 18),
    "0xe4f05A66Ec68B54A58B17c22107b02e0232cC817": ("WKLAY", 18),
    "0x5fff3a6c16c2208103f318f4713d4d90601a7313": ("KDAI", 18),
    "0x02cBe46fB8A1F579254a9B485788f2D86Cad51aa": ("bKLAY", 18),
    "0xDCbacF3f7a069922E677912998c8d57423C37dfA": ("WEMIX$", 18),
    "0x588C62eD9aa7367d7cd9C2A9aaAc77e44fe8221B": ("KSP", 18),
    "0x275f942985503d8CE9558f8377CC526DcD3d80b0": ("KUSDT", 6),
    "0x210BC03F49052169D5588A52C317f71cF2078b85": ("KUSDC", 6),
    "0xDd483a970a7A7FeF2B223403f18a1548DAc5FbC7": ("KETH", 18),
    "0xd6dAb4CfF47dF175349e6e7eE2BF7c40Bb8c05A3": ("KXRP", 6),
    "0x321BC0B63EFb1e4af08ec6d20C85d5E94dDDc3C8": ("MOOI", 18),
    "0x46f307b58bf05Ff089BA23799FAE0e518557f87c": ("MBX", 18),
    "0x9EAeFb09fe4AFDDd797a1F78E3c9aFB39fD0B73b": ("BOA", 7),
    "0x8e130Dc1112D5B9fDFd9F4D75F0E63E9ff17f964": ("SSX", 18),
    "0x27DcD181459bCDdc63C37BAb1E404A313C0dfD79": ("BTT", 18),
    "0x1BcFc0B5090e8c0f2C9eCA6afDc8a4F59E54f5B4": ("MBL", 18),
    "0x97bf7Fd81E6Cb0E3C8E42dFDad6a5e8B113C2a1f": ("PER", 18),
    "0x74BA03198FEd2b15a51AF242b9c63FAF3C8f4D34": ("FNSA", 18),
    "0x4B91a382FA6C02B16E57b1f798a9F01E25bA88c3": ("TEMCO", 18),
    "0x119883ee408aA5B9f891979fe73E69f8A4B45DA7": ("MIX", 18),
    # Korean exchange tokens
    "0x275f942985503d8ce9558f8377cc526a3aba3566": ("WIKEN", 18),
    "0xafde910130c335fa5bd5fe991053e3e0a49dce7b": ("PIB", 18),
    "0xa7b4c080c7b9815980a7fb7c4b2bf1e021609cba": ("LTR", 18),
    "0x472fac08cf4836bee54343edfb49023746b27933": ("SPIN", 18),
    "0xd364de0683b29e582e5713425b215b24ce804ae9": ("CENT", 18),
    "0xe993e5668a034a98cc53cc1e3bfba910119440a1": ("CELL", 18),
    "0xe3ecbfbb8f8c37c2450b9920b79755e5280a4252": ("SNTC", 18),
    "0x96035fbdd4cb888862ee28c9d8fdadef78311cc9": ("MM", 18),
}
def get_kaia(addr):
    rpc = os.getenv("KAIA_RPC_URL", KAIA_RPC)
    resp = _evm_rpc(rpc, "eth_getBalance", [addr, "latest"])
    raw = hex2int(resp["result"])
    results = [R("KAIA", addr, "KAIA", fmt(raw, 18), raw, 18, None, "native")]
    # Kaiascan API (키 있을 때만)
    kk = os.getenv("KAIASCAN_API_KEY", "").strip()
    token_found = False
    if kk:
        hdr = {"accept": "application/json", "x-api-key": kk}
        page = 1
        while True:
            try:
                data = http_get(
                    "https://mainnet-oapi.kaiascan.io/api/v1/accounts/fungible-token-balances",
                    headers=hdr,
                    params={"accountAddress": addr, "page": page, "size": 100}
                )
            except Exception:
                break
            items = (data.get("result", {}).get("items")
                     or data.get("result", {}).get("results")
                     or data.get("items") or data.get("results") or [])
            if not items:
                break
            token_found = True
            for it in items:
                r = str(it.get("balance") or it.get("token_balance") or "0")
                if r == "0": continue
                sym = it.get("token_symbol") or it.get("symbol") or "UNKNOWN"
                d = it.get("token_decimal") or it.get("decimals")
                ct = it.get("contract_address") or it.get("token_address")
                if d is None:
                    amt, dv = r, None
                else:
                    dv = int(d)
                    amt = fmt(r, dv)
                results.append(R("KAIA", addr, sym, amt, r, dv, ct))
            if len(items) < 100: break
            page += 1
    # Kaiascan 키 없을 때 → batch RPC로 알려진 토큰 체크 (빠름)
    if not token_found:
        addr_padded = "0" * 24 + addr[2:].lower()
        token_list = list(KAIA_KNOWN_TOKENS.items())
        calls = []
        for i, (ca, (sym, dec)) in enumerate(token_list):
            calls.append({
                "jsonrpc": "2.0", "id": i + 1,
                "method": "eth_call",
                "params": [{"to": ca, "data": "0x70a08231" + addr_padded}, "latest"]
            })
        try:
            resp_list = _evm_batch_rpc(rpc, calls)
            resp_map = {r2["id"]: r2 for r2 in resp_list}
            for i, (ca, (sym, dec)) in enumerate(token_list):
                r2 = resp_map.get(i + 1, {})
                hex_val = r2.get("result", "0x0")
                if not hex_val or hex_val in ("0x", "0x0", "0x" + "0" * 64):
                    continue
                rv = str(int(hex_val, 16))
                if rv == "0":
                    continue
                results.append(R("KAIA", addr, sym, fmt(rv, dec), rv, dec, ca))
        except Exception:
            pass
    return results
# ═══════════════════════════════════════════════
#  SGB (Songbird) — 공개 RPC
# ═══════════════════════════════════════════════
def get_sgb(addr):
    resp = _evm_rpc("https://songbird-api.flare.network/ext/C/rpc",
                    "eth_getBalance", [addr, "latest"])
    raw = hex2int(resp["result"])
    return [R("SGB", addr, "SGB", fmt(raw, 18), raw, 18, None, "native")]
# ═══════════════════════════════════════════════
#  XRP  — 공식 Ripple RPC (키 불필요)
#  40자 hex currency code 디코딩 지원
# ═══════════════════════════════════════════════
def _xrp_decode_currency(code: str) -> str:
    """XRPL currency code 디코딩: 3자=그대로, 40자hex=디코딩"""
    if not code:
        return "UNKNOWN"
    if len(code) == 3:
        return code
    if len(code) == 40:
        try:
            decoded = bytes.fromhex(code).decode("ascii", errors="ignore").strip("\x00").strip()
            if decoded and decoded.isprintable():
                return decoded
        except Exception:
            pass
    return code
def get_xrp(addr):
    rpc = os.getenv("XRP_RPC_URL", "https://s1.ripple.com:51234/")
    info = http_post(rpc, {"method": "account_info",
                           "params": [{"account": addr, "ledger_index": "validated"}]})
    lines = http_post(rpc, {"method": "account_lines",
                            "params": [{"account": addr, "ledger_index": "validated"}]})
    raw_drops = str(info["result"]["account_data"]["Balance"])
    results = [R("XRP", addr, "XRP", fmt(raw_drops, 6), raw_drops, 6, None, "native")]
    for ln in lines["result"].get("lines", []):
        bal = noexp(str(ln["balance"]))
        cur_raw = ln.get("currency", "UNKNOWN")
        cur = _xrp_decode_currency(cur_raw)
        issuer = ln.get("account")
        if bal not in ("0", "0.0", "0.000000", ""):
            results.append(R("XRP", addr, cur, bal, None, None, issuer,
                             "token", issuer=issuer))
    return results
# ═══════════════════════════════════════════════
#  XLM  — Horizon API (키 불필요)
# ═══════════════════════════════════════════════
def get_xlm(addr):
    data = http_get(f"https://horizon.stellar.org/accounts/{addr}")
    results = []
    for b in data.get("balances", []):
        bal = noexp(str(b.get("balance", "0")))
        if b.get("asset_type") == "native":
            results.append(R("XLM", addr, "XLM", bal, None, 7, None, "native"))
        else:
            code = b.get("asset_code", "UNKNOWN")
            iss = b.get("asset_issuer")
            results.append(R("XLM", addr, code, bal, None, None, iss,
                             "token", issuer=iss))
    return results
# ═══════════════════════════════════════════════
#  SOL  — 공개 RPC (키 불필요)
# ═══════════════════════════════════════════════
def _sol(method, params):
    rpc = os.getenv("SOL_RPC_URL", "https://api.mainnet-beta.solana.com")
    return http_post(rpc, {"jsonrpc": "2.0", "id": 1,
                           "method": method, "params": params})
def get_sol(addr):
    native = _sol("getBalance", [addr])
    tokens = _sol("getTokenAccountsByOwner", [
        addr,
        {"programId": "TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA"},
        {"encoding": "jsonParsed"}
    ])
    raw_l = str(native["result"]["value"])
    results = [R("SOL", addr, "SOL", fmt(raw_l, 9), raw_l, 9, None, "native")]
    for it in tokens["result"]["value"]:
        info = it["account"]["data"]["parsed"]["info"]
        ta = info["tokenAmount"]
        raw = str(ta["amount"])
        if raw == "0": continue
        dec = int(ta["decimals"])
        mint = info["mint"]
        results.append(R("SOL", addr, mint, fmt(raw, dec), raw, dec, mint))
    return results
# ═══════════════════════════════════════════════
#  TRX  — TronGrid (키 없어도 동작, 있으면 rate limit ↑)
# ═══════════════════════════════════════════════
def _trx_hdr():
    h = {}
    k = os.getenv("TRONGRID_API_KEY", "").strip()
    if k: h["TRON-PRO-API-KEY"] = k
    return h
def get_trx(addr):
    hdr = _trx_hdr()
    # native TRX
    acct = http_post("https://api.trongrid.io/wallet/getaccount",
                     {"address": addr, "visible": True}, headers=hdr)
    raw_sun = str(acct.get("balance", 0))
    results = [R("TRX", addr, "TRX", fmt(raw_sun, 6), raw_sun, 6, None, "native")]
    # TronScan token_asset_overview — 토큰명, decimals, 잔고를 한 번에 조회
    try:
        data = http_get(
            "https://apilist.tronscanapi.com/api/account/token_asset_overview",
            params={"address": addr}
        )
        for tok in data.get("data", []):
            sym = tok.get("tokenAbbr") or tok.get("tokenName") or "UNKNOWN"
            dec = int(tok.get("tokenDecimal", 0) or 0)
            raw_bal = str(tok.get("balance", "0"))
            ttype = tok.get("tokenType", "")
            ct = tok.get("tokenId") or tok.get("contractAddress") or ""
            # native TRX는 위에서 이미 처리
            if sym.lower() == "trx" and ttype in ("trc10", ""):
                continue
            if raw_bal == "0":
                continue
            results.append(R("TRX", addr, sym, fmt(raw_bal, dec), raw_bal, dec, ct,
                             "native" if ttype == "" else "token"))
    except Exception:
        # TronScan 실패 시 TronGrid fallback (토큰명 제한적)
        try:
            tg = http_get(f"https://api.trongrid.io/v1/accounts/{addr}", headers=hdr)
            for token_map in tg.get("data", [{}])[0].get("trc20", []):
                for ca, rb in token_map.items():
                    rs = str(rb)
                    if rs == "0": continue
                    # 개별 contract 조회 대신 주소 축약 표시
                    results.append(R("TRX", addr, ca[:8] + "…", rs, rs, 0, ca))
        except Exception:
            pass
    return results
# ═══════════════════════════════════════════════
#  ONT  — 공식 노드 다중 폴백 (키 불필요)
#  ONT: decimals=9, ONG: decimals=18
# ═══════════════════════════════════════════════
ONT_RPC_ENDPOINTS = [
    "https://dappnode1.ont.io:10334",
    "http://dappnode1.ont.io:20334",
    "http://dappnode2.ont.io:20334",
]
def get_ont(addr):
    # 1) native (ONT, ONG) via RPC
    data = None
    for ep in ONT_RPC_ENDPOINTS:
        try:
            data = http_get(f"{ep}/api/v1/balancev2/{addr}")
            break
        except Exception:
            continue
    results = []
    if data:
        result = data.get("Result") or data.get("result") or {}
        if "ont" in result:
            raw_ont = str(result["ont"])
            results.append(R("ONT", addr, "ONT", fmt(raw_ont, 9), raw_ont, 9, None, "native"))
        if "ong" in result:
            raw_ong = str(result["ong"])
            results.append(R("ONT", addr, "ONG", fmt(raw_ong, 18), raw_ong, 18, None, "token"))
    # 2) explorer.ont.io v2 — native + OEP4 + ORC20 토큰
    try:
        # native 폴백 (RPC 실패 시)
        if not results:
            nd = http_get(f"https://explorer.ont.io/v2/addresses/{addr}/native/balances")
            for item in nd.get("result", []):
                name = item.get("asset_name", "")
                bal = noexp(str(item.get("balance", "0")))
                if name == "ont" and bal not in ("0", "0.0"):
                    results.append(R("ONT", addr, "ONT", bal, None, 9, None, "native"))
                elif name == "ong" and bal not in ("0", "0.0"):
                    results.append(R("ONT", addr, "ONG", bal, None, 18, None, "token"))
        # OEP4 토큰
        oep4 = http_get(f"https://explorer.ont.io/v2/addresses/{addr}/oep4/balances")
        for item in oep4.get("result", []):
            bal = noexp(str(item.get("balance", "0")))
            if bal in ("0", "0.0", ""):
                continue
            sym = item.get("asset_name") or "UNKNOWN"
            ct = item.get("contract_hash")
            results.append(R("ONT", addr, sym, bal, None, None, ct))
        # ORC20 토큰
        orc20 = http_get(f"https://explorer.ont.io/v2/addresses/{addr}/orc20/balances")
        for item in orc20.get("result", []):
            bal = noexp(str(item.get("balance", "0")))
            if bal in ("0", "0.0", ""):
                continue
            sym = item.get("asset_name") or "UNKNOWN"
            ct = item.get("contract_hash")
            results.append(R("ONT", addr, sym, bal, None, None, ct))
    except Exception:
        pass
    if not results:
        raise RuntimeError("ONT 잔고 조회 실패 — 노드 및 Explorer 응답 없음")
    return results
# ═══════════════════════════════════════════════
#  ARDR  — Jelurida 공개 노드 (키 불필요)
# ═══════════════════════════════════════════════
def _ardor(params):
    node = os.getenv("ARDOR_NODE_URL", "https://ardor.jelurida.com/nxt")
    return http_get(node, params=params)
def get_ardr(addr):
    native = _ardor({"requestType": "getBalance", "chain": "1", "account": addr})
    assets = _ardor({"requestType": "getAccountAssets",
                     "account": addr, "includeAssetInfo": "true"})
    raw_nqt = str(native.get("balanceNQT", "0"))
    results = [R("ARDR", addr, "ARDR", fmt(raw_nqt, 8), raw_nqt, 8, None, "native")]
    # IGNIS (child chain 2) 잔고도 함께 표시
    try:
        ignis = _ardor({"requestType": "getBalance", "chain": "2", "account": addr})
        ignis_raw = str(ignis.get("balanceNQT", "0"))
        if ignis_raw != "0":
            results.append(R("ARDR", addr, "IGNIS", fmt(ignis_raw, 8), ignis_raw, 8, None, "token"))
    except Exception:
        pass
    for a in assets.get("accountAssets", []):
        rq = str(a.get("quantityQNT", "0"))
        if rq == "0": continue
        d = int(a.get("decimals", 0) or 0)
        sym = a.get("name") or a.get("asset") or "ASSET"
        results.append(R("ARDR", addr, sym, fmt(rq, d), rq, d, a.get("asset")))
    return results
# ═══════════════════════════════════════════════
#  IGNIS  — Ardor child chain 2 (키 불필요)
# ═══════════════════════════════════════════════
def get_ignis(addr):
    native = _ardor({"requestType": "getBalance", "chain": "2", "account": addr})
    raw_nqt = str(native.get("balanceNQT", "0"))
    results = [R("IGNIS", addr, "IGNIS", fmt(raw_nqt, 8), raw_nqt, 8, None, "native")]
    # ARDR 잔고도 참고로 표시
    try:
        ardr = _ardor({"requestType": "getBalance", "chain": "1", "account": addr})
        ardr_raw = str(ardr.get("balanceNQT", "0"))
        if ardr_raw != "0":
            results.append(R("IGNIS", addr, "ARDR", fmt(ardr_raw, 8), ardr_raw, 8, None, "token"))
    except Exception:
        pass
    return results
# ═══════════════════════════════════════════════
#  EOS  — Hyperion (공개 엔드포인트 기본 내장)
# ═══════════════════════════════════════════════
def get_eos(addr):
    hyperion = os.getenv("EOS_HYPERION_URL", "https://eos.hyperion.eosrio.io")
    url = f"{hyperion.rstrip('/')}/v2/state/get_tokens"
    data = http_get(url, params={"account": addr})
    tokens = data.get("tokens") if isinstance(data, dict) else None
    if tokens is None:
        tokens = data.get("results") if isinstance(data, dict) else None
    if tokens is None and isinstance(data, list):
        tokens = data
    if not tokens:
        tokens = []
    results = []
    for it in tokens:
        sym = it.get("symbol") or "UNKNOWN"
        ct = it.get("contract") or it.get("account") or it.get("code")
        # amount는 숫자(float/int) 또는 "123.4567 EOS" 형태 문자열
        raw_amount = it.get("amount")
        if raw_amount is None:
            raw_amount = it.get("balance", 0)
        if isinstance(raw_amount, str) and " " in raw_amount:
            amt_str, sym = raw_amount.split(" ", 1)
        elif isinstance(raw_amount, (int, float)):
            amt_str = noexp(str(raw_amount))
        else:
            amt_str = noexp(str(raw_amount))
        # precision 정보 활용
        prec = it.get("precision")
        if prec is not None:
            # precision에 맞게 소수점 포맷
            amt_str = f"{float(amt_str):.{int(prec)}f}"
        dec = len(amt_str.split(".")[1]) if "." in amt_str else 0
        results.append(R("EOS", addr, sym, amt_str, None, dec, ct,
                         "native" if sym == "EOS" else "token"))
    return results
# ═══════════════════════════════════════════════
#  PCI (Paycoin) — scan.payprotocol.io
# ═══════════════════════════════════════════════
def get_pci(addr):
    url = f"https://scan.payprotocol.io/api/account/{addr}"
    # 1) 일반 요청
    try:
        data = http_get(url)
        if isinstance(data, dict):
            return [R("PCI", addr, "PCI", noexp(str(data.get("balance", "0"))),
                       None, None, None, "native")]
    except Exception:
        pass
    # 2) SSL 인증서 오류 우회 재시도
    try:
        r = requests.get(url, timeout=TIMEOUT, verify=False,
                         headers={"User-Agent": "multichain-balance-checker/5.0"})
        r.raise_for_status()
        data = r.json()
        if isinstance(data, dict):
            return [R("PCI", addr, "PCI", noexp(str(data.get("balance", "0"))),
                       None, None, None, "native")]
    except Exception:
        pass
    raise RuntimeError(
        "PCI 잔고 조회 실패 — scan.payprotocol.io 서비스 상태를 확인하세요 (현재 다운 가능)")
# ═══════════════════════════════════════════════
#  체인 목록
# ═══════════════════════════════════════════════
CHAINS = {
    "BTC":   ("Bitcoin",       get_btc),
    "LTC":   ("Litecoin",      get_ltc),
    "BCH":   ("Bitcoin Cash",  get_bch),
    "RVN":   ("Ravencoin",     get_rvn),
    "ETH":   ("Ethereum",      lambda a: get_evm_balances("ETH", a)),
    "BNB":   ("BNB Chain",     lambda a: get_evm_balances("BNB", a)),
    "POL":   ("Polygon",       lambda a: get_evm_balances("POL", a)),
    "KAIA":  ("Kaia",          get_kaia),
    "SGB":   ("Songbird",      get_sgb),
    "XRP":   ("XRP Ledger",    get_xrp),
    "XLM":   ("Stellar",       get_xlm),
    "SOL":   ("Solana",        get_sol),
    "TRX":   ("Tron",          get_trx),
    "ONT":   ("Ontology",      get_ont),
    "ARDR":  ("Ardor",         get_ardr),
    "IGNIS": ("Ignis",         get_ignis),
    "EOS":   ("EOS",           get_eos),
    "PCI":   ("Paycoin",       get_pci),
}
def get_balances(chain: str, address: str) -> dict:
    chain = chain.upper().strip()
    if chain not in CHAINS:
        raise ValueError(f"미지원 체인: {chain}")
    name, fn = CHAINS[chain]
    return {"chain": chain, "name": name, "address": address, "assets": fn(address)}
# ═══════════════════════════════════════════════
#  대화형 CLI
# ═══════════════════════════════════════════════
def print_assets(result):
    chain = result["chain"]
    name = result.get("name", "")
    address = result["address"]
    assets = result.get("assets", [])
    print(f"\n{'='*60}")
    print(f"  {chain} ({name})  |  {address}")
    print(f"{'='*60}")
    if not assets:
        print("  (잔고 없음)")
        return
    mx = min(max(len(a["ticker"]) for a in assets), 20)
    for a in assets:
        tk = a["ticker"]
        if len(tk) > 20:
            tk = tk[:8] + "..." + tk[-6:]
        tag = "[native]" if a["type"] == "native" else "[token] "
        print(f"  {tag}  {tk:>{mx}s}  {a['amount']}")
    print()
def _save_results_xlsx(all_results, out_path=None):
    """조회 결과 리스트를 .xlsx 파일로 저장"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("  [오류] openpyxl 패키지 필요: pip install openpyxl")
        return
    if not all_results:
        return
    if out_path is None:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = f"잔고조회_{ts}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "잔고 조회 결과"
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    headers = ["체인", "체인명", "주소", "타입", "티커", "잔고", "컨트랙트"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row_num = 2
    for result in all_results:
        if "error" in result and not result.get("assets"):
            ws.cell(row=row_num, column=1, value=result.get("chain", ""))
            ws.cell(row=row_num, column=3, value=result.get("address", ""))
            ws.cell(row=row_num, column=6, value=f"[오류] {result['error']}")
            for c in range(1, 8):
                ws.cell(row=row_num, column=c).border = thin_border
            row_num += 1
            continue
        chain = result["chain"]
        name = result.get("name", "")
        address = result["address"]
        for a in result.get("assets", []):
            ws.cell(row=row_num, column=1, value=chain).border = thin_border
            ws.cell(row=row_num, column=2, value=name).border = thin_border
            ws.cell(row=row_num, column=3, value=address).border = thin_border
            atype = "native" if a["type"] == "native" else "token"
            ws.cell(row=row_num, column=4, value=atype).border = thin_border
            ws.cell(row=row_num, column=5, value=a["ticker"]).border = thin_border
            amt_cell = ws.cell(row=row_num, column=6, value=a["amount"])
            amt_cell.alignment = Alignment(horizontal="right")
            amt_cell.border = thin_border
            ws.cell(row=row_num, column=7, value=a.get("contract") or "").border = thin_border
            row_num += 1
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 48
    wb.save(out_path)
    print(f"\n  >> 엑셀 저장 완료: {out_path}  ({row_num - 2}행)\n")
def _parse_pairs(lines):
    """텍스트 줄들에서 (chain, address) 쌍 추출, 중복 제거"""
    pairs = []
    seen = set()
    for line in lines:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.replace(",", " ").replace("\t", " ").split()
        if len(parts) < 2:
            print(f"  [건너뜀] 형식 오류: {line}")
            continue
        chain = parts[0].upper().strip()
        address = parts[1].strip()
        if chain not in CHAINS:
            print(f"  [건너뜀] 미지원 체인: {chain}")
            continue
        key = (chain, address.lower())
        if key in seen:
            continue
        seen.add(key)
        pairs.append((chain, address))
    return pairs
def batch_text():
    """텍스트 일괄 입력 모드"""
    cl = list(CHAINS.keys())
    print("\n┌─ 지원 체인 ─────────────────────────────────┐")
    for i, ck in enumerate(cl, 1):
        print(f"│  {ck:<6s} ({CHAINS[ck][0]})")
    print("└─────────────────────────────────────────────┘")
    print("\n[배치 입력] 체인 주소를 한 줄씩 입력하세요.")
    print("  형식: CHAIN ADDRESS  (예: ETH 0xabc...)")
    print("  빈 줄 입력하면 조회 시작\n")
    lines = []
    while True:
        try:
            line = input("  > ").strip()
        except EOFError:
            break
        if not line:
            break
        lines.append(line)
    pairs = _parse_pairs(lines)
    if not pairs:
        print("  조회할 항목이 없습니다.")
        return []
    print(f"\n총 {len(pairs)}개 조회 시작...\n")
    all_results = []
    for i, (chain, address) in enumerate(pairs, 1):
        print(f"[{i}/{len(pairs)}] {chain} {address[:20]}...")
        try:
            result = get_balances(chain, address)
            print_assets(result)
            all_results.append(result)
        except Exception as e:
            print(f"  [오류] {e}\n")
            all_results.append({"chain": chain, "address": address, "error": str(e)})
    _save_results_xlsx(all_results)
    return all_results
def batch_excel(filepath):
    """엑셀 파일 입력 → 조회 → 엑셀 파일 출력"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  [오류] openpyxl 패키지가 필요합니다: pip install openpyxl")
        return
    if not os.path.isfile(filepath):
        print(f"  [오류] 파일을 찾을 수 없습니다: {filepath}")
        return
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    lines = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = [str(c).strip() if c else "" for c in row]
        if len(cells) >= 2 and cells[0] and cells[1]:
            chain_val = cells[0].upper().strip()
            if chain_val in ("CHAIN", "체인", "메인넷", "MAINNET", "NETWORK"):
                continue
            lines.append(f"{cells[0]} {cells[1]}")
    wb.close()
    pairs = _parse_pairs(lines)
    if not pairs:
        print("  엑셀에서 유효한 항목을 찾지 못했습니다.")
        print("  형식: A열=체인(ETH,BTC,...), B열=주소")
        return
    print(f"\n엑셀에서 {len(pairs)}개 항목 로드 완료. 조회 시작...\n")
    all_results = []
    for i, (chain, address) in enumerate(pairs, 1):
        print(f"[{i}/{len(pairs)}] {chain} {address[:20]}...")
        try:
            result = get_balances(chain, address)
            print_assets(result)
            all_results.append(result)
        except Exception as e:
            print(f"  [오류] {e}\n")
            all_results.append({"chain": chain, "address": address, "error": str(e), "assets": []})
    base = os.path.splitext(filepath)[0]
    _save_results_xlsx(all_results, f"{base}_결과.xlsx")
def interactive():
    cl = list(CHAINS.keys())
    all_results = []
    while True:
        print("\n┌─ 지원 체인 (API 키 불필요) ──────────────────┐")
        for i, ck in enumerate(cl, 1):
            print(f"│  {i:2d}. {ck:<6s} ({CHAINS[ck][0]})")
        print("├─────────────────────────────────────────────┤")
        print("│   B. 배치 입력 (여러 체인+주소 한번에)")
        print("│   X. 엑셀 파일 입력/출력 (.xlsx)")
        print("│   0. 종료")
        print("└─────────────────────────────────────────────┘")
        sel = input("선택: ").strip()
        if sel == "0" or sel.lower() in ("q", "quit", "exit"):
            break
        if sel.upper() == "B":
            batch_text()
            continue
        if sel.upper() == "X":
            fp = input("엑셀 파일 경로 (.xlsx): ").strip()
            if fp:
                batch_excel(fp)
            continue
        if sel.isdigit():
            idx = int(sel) - 1
            if not (0 <= idx < len(cl)):
                print("잘못된 번호"); continue
            chain = cl[idx]
        elif sel.upper() in CHAINS:
            chain = sel.upper()
        else:
            print("잘못된 입력"); continue
        address = input(f"{chain} 주소 입력: ").strip()
        if not address:
            print("주소 비어있음"); continue
        try:
            result = get_balances(chain, address)
            print_assets(result)
            all_results.append(result)
        except Exception as e:
            print(f"\n  [오류] {e}\n")
        if input("계속? (Enter=계속 / q=종료): ").strip().lower() in ("q", "quit", "exit"):
            break
    if all_results:
        _save_results_xlsx(all_results)
if __name__ == "__main__":
    if len(sys.argv) == 3:
        try:
            print(json.dumps(get_balances(sys.argv[1], sys.argv[2]),
                             ensure_ascii=False, indent=2))
        except Exception as e:
            print(json.dumps({"error": str(e)}, ensure_ascii=False, indent=2))
    elif len(sys.argv) == 2 and sys.argv[1].endswith(".xlsx"):
        batch_excel(sys.argv[1])
    else:
        interactive()
